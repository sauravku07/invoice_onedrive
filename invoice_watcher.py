import os
import re
import time
import shutil
import pandas as pd
import pdfplumber
import pytesseract
from PIL import Image
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ================== PATHS ==================
ONEDRIVE = os.environ.get("OneDrive")
if not ONEDRIVE:
    raise RuntimeError("❌ OneDrive not found")

BASE = os.path.join(ONEDRIVE, "Invoices")
INPUT = os.path.join(BASE, "Input")
PROCESSED = os.path.join(BASE, "Processed")
EXCEL = os.path.join(BASE, "Invoice_Data.xlsx")

os.makedirs(INPUT, exist_ok=True)
os.makedirs(PROCESSED, exist_ok=True)

HEADERS = [
    "Sr.No", "Invoice Date", "Invoice No", "Ref No",
    "Particular", "Amount", "TDS (10%)",
    "Clear Amount", "Comment"
]

# ================== SAFE EXCEL ==================
def wait_for_excel():
    while True:
        try:
            with open(EXCEL, "a"):
                return
        except PermissionError:
            print("⏳ Excel open, waiting...")
            time.sleep(2)

def load_excel():
    if not os.path.exists(EXCEL):
        df = pd.DataFrame(columns=HEADERS)
        df.to_excel(EXCEL, index=False)
        apply_formulas_and_formatting()
        return df

    wait_for_excel()
    df = pd.read_excel(EXCEL)
    df = df.reindex(columns=HEADERS)
    return df

def save_excel(df):
    wait_for_excel()
    df.to_excel(EXCEL, index=False)
    apply_formulas_and_formatting()

# ================== FORMATTING + FORMULAS ==================
def apply_formulas_and_formatting():
    wb = load_workbook(EXCEL)
    ws = wb.active

    YELLOW = PatternFill("solid", fgColor="FFFFF200")
    BOLD = Font(bold=True)

    # Header
    for cell in ws[1]:
        cell.fill = YELLOW
        cell.font = BOLD

    # Remove old TOTAL
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 5).value == "TOTAL":
            ws.delete_rows(r)

    last_row = ws.max_row
    total_row = last_row + 1

    # TOTAL formulas
    ws.cell(total_row, 5).value = "TOTAL"
    ws.cell(total_row, 6).value = f"=SUM(F2:F{last_row})"
    ws.cell(total_row, 7).value = f"=SUM(G2:G{last_row})"

    for col in range(1, 10):
        ws.cell(total_row, col).fill = YELLOW
        ws.cell(total_row, col).font = BOLD

    # AUTO COLUMN WIDTH
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 3

    wb.save(EXCEL)

# ================== OCR ==================
def ocr_file(path):
    text = ""
    if path.lower().endswith(".pdf"):
        with pdfplumber.open(path) as pdf:
            for p in pdf.pages:
                if p.extract_text():
                    text += p.extract_text() + " "
    else:
        text = pytesseract.image_to_string(Image.open(path))
    return re.sub(r"\s+", " ", text)

# ================== PARTICULAR ==================
def extract_particular(text):
    case_types = [
        r"Writ Petition\s*\(C\)", r"CS\s*\(COMM\)", r"LPA",
        r"IPD", r"FAO", r"RFA", r"CM", r"ARB\.?P", r"OMP"
    ]

    pattern = re.compile(
        rf"({'|'.join(case_types)})\s*No\.?\s*(\d+)\s*of\s*(\d{{4}}).*?"
        r"before\s+the\s+([A-Za-z\s]+Court(?:\s+at\s+[A-Za-z\s]+)?)",
        re.I
    )

    matches = pattern.findall(text)
    if not matches:
        return ""

    grouped = defaultdict(list)
    for case, num, year, court in matches:
        grouped[(case.upper(), year, court.strip())].append(int(num))

    results = []
    for (case, year, court), nums in grouped.items():
        nums = sorted(set(nums))
        ranges = []
        s = p = nums[0]

        for n in nums[1:]:
            if n == p + 1:
                p = n
            else:
                ranges.append(f"{s}-{p}" if s != p else str(s))
                s = p = n
        ranges.append(f"{s}-{p}" if s != p else str(s))

        results.append(
            f"{case.title()} No. {', '.join(ranges)} of {year} before the {court}"
        )

    return "; ".join(results)

# ================== AMOUNT ==================
def extract_amount(text):
    for p in [
        r"Total\s*Invoice\s*Value.*?([0-9,]+\.\d{2})",
        r"Grand\s*Total.*?([0-9,]+\.\d{2})",
        r"Total\s*Amount.*?([0-9,]+\.\d{2})"
    ]:
        m = re.search(p, text, re.I)
        if m:
            return float(m.group(1).replace(",", ""))
    return 0.0

# ================== PROCESS FILE ==================
def process_file(path):
    time.sleep(2)
    text = ocr_file(path)

    invs = re.findall(r"\b[A-Z0-9]{6,}\b", text)
    if not invs:
        print("⚠ Invoice No not found")
        return

    inv = max(invs, key=len)
    df = load_excel()

    if inv in df["Invoice No"].astype(str).values:
        print("⚠ Duplicate skipped:", inv)
        shutil.move(path, os.path.join(PROCESSED, os.path.basename(path)))
        return

    date = re.search(r"\d{1,2}-[A-Za-z]{3}-\d{4}", text)
    ref = re.search(r"(Our\s*Ref|Ref)\s*[:\-]?\s*([A-Z0-9\/\-]+)", text, re.I)

    amt = extract_amount(text)
    df = df[df["Particular"] != "TOTAL"]

    df.loc[len(df)] = [
        len(df) + 1,
        date.group() if date else "",
        inv,
        ref.group(2) if ref else "",
        extract_particular(text),
        amt,
        round(amt * 0.10, 2),
        round(amt * 0.90, 2),
        ""
    ]

    save_excel(df)
    shutil.move(path, os.path.join(PROCESSED, os.path.basename(path)))
    print("✔ Processed:", os.path.basename(path))

# ================== WATCHDOG ==================
class Handler(FileSystemEventHandler):
    def on_created(self, e):
        if not e.is_directory:
            process_file(e.src_path)

# ================== MAIN ==================
if __name__ == "__main__":
    print("🚀 Invoice Automation Running")

    for f in os.listdir(INPUT):
        p = os.path.join(INPUT, f)
        if os.path.isfile(p):
            process_file(p)

    obs = Observer()
    obs.schedule(Handler(), INPUT, recursive=False)
    obs.start()

    try:
        while True:
            time.sleep(5)
    except KeyboardInterrupt:
        obs.stop()

    obs.join()



